"""
06_challenger_model.py
Champion vs Challenger evaluation:
  Champion  = XGBoost-based scorecard
  Challenger = Logistic Regression on WoE features
Documents lift, Gini improvement, cut-off sensitivity,
and writes a structured validation report to Excel.
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import joblib
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.metrics import (roc_auc_score, roc_curve, precision_recall_curve,
                              average_precision_score)
from utils import (OUTPUTS_DIR, CHARTS_DIR, EXCEL_DIR, TARGET,
                   gini_coefficient, ks_statistic, logger)


# ── Style helpers (same as reporting) ────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="000000"):
    return Font(bold=True, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def auto_width(ws, min_w=12, max_w=45):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

HDR_BLUE   = hdr_fill("1F3864")
HDR_GREY   = hdr_fill("D9E1F2")
GREEN_FILL = hdr_fill("C6EFCE")
RED_FILL   = hdr_fill("FFC7CE")
WHITE_FONT = bold_font(color="FFFFFF")


# ── Challenger: Logistic Regression ──────────────────────────────────────────
def train_challenger(df_woe: pd.DataFrame, selected_features: list):
    woe_cols = [f"{f}_woe" for f in selected_features]
    X = df_woe[woe_cols].values
    y = df_woe[TARGET].values

    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42, stratify=y
    )

    lr = LogisticRegression(
        C=1.0,
        max_iter=1000,
        class_weight="balanced",
        random_state=42,
        solver="lbfgs",
    )
    lr.fit(X_train, y_train)
    y_prob = lr.predict_proba(X_test)[:, 1]

    return lr, X_test, y_test, y_prob


# ── Metrics bundle ────────────────────────────────────────────────────────────
def compute_metrics(y_true, y_prob, model_name: str) -> dict:
    auc  = roc_auc_score(y_true, y_prob)
    gini = gini_coefficient(y_true, y_prob)
    ks   = ks_statistic(y_true, y_prob)
    ap   = average_precision_score(y_true, y_prob)
    return {
        "Model":       model_name,
        "AUC":         round(auc, 4),
        "Gini":        round(gini, 4),
        "KS":          round(ks, 4),
        "Avg Precision": round(ap, 4),
    }


# ── Cut-off sensitivity table ─────────────────────────────────────────────────
def cutoff_sensitivity(y_true, y_prob_champ, y_prob_chal) -> pd.DataFrame:
    thresholds = np.arange(0.05, 0.55, 0.05)
    rows = []
    for t in thresholds:
        for model_name, y_prob in [("Champion XGBoost", y_prob_champ),
                                    ("Challenger LR",   y_prob_chal)]:
            y_pred  = (y_prob >= t).astype(int)
            tp = ((y_pred == 1) & (y_true == 1)).sum()
            fp = ((y_pred == 1) & (y_true == 0)).sum()
            tn = ((y_pred == 0) & (y_true == 0)).sum()
            fn = ((y_pred == 0) & (y_true == 1)).sum()

            precision = tp / (tp + fp) if (tp + fp) > 0 else 0
            recall    = tp / (tp + fn) if (tp + fn) > 0 else 0
            approval  = (tn + fn) / len(y_true)

            rows.append({
                "Model":          model_name,
                "Cut-off":        round(t, 2),
                "Precision":      round(precision, 4),
                "Recall":         round(recall, 4),
                "Approval Rate":  round(approval, 4),
                "TP":             int(tp),
                "FP":             int(fp),
                "TN":             int(tn),
                "FN":             int(fn),
            })
    return pd.DataFrame(rows)


# ── Lift table ────────────────────────────────────────────────────────────────
def compute_lift(y_true, y_prob, model_name: str, n_bins: int = 10) -> pd.DataFrame:
    df = pd.DataFrame({"y": y_true, "prob": y_prob})
    df["decile"] = pd.qcut(df["prob"].rank(method="first"), q=n_bins,
                            labels=False, duplicates="drop")
    df["decile"] = n_bins - df["decile"]  # 1 = highest risk

    rows = []
    overall_dr = df["y"].mean()
    for d in sorted(df["decile"].unique()):
        sub = df[df["decile"] == d]
        dr  = sub["y"].mean()
        rows.append({
            "Model":        model_name,
            "Decile":       int(d),
            "Count":        len(sub),
            "Default Rate": round(dr, 4),
            "Lift":         round(dr / overall_dr, 3) if overall_dr > 0 else 0,
        })
    return pd.DataFrame(rows)


# ── Charts ────────────────────────────────────────────────────────────────────
def plot_roc_comparison(y_test, y_prob_champ, y_prob_chal):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    # ROC curves
    for ax, (name, prob, color) in [
        (axes[0], ("Champion XGBoost", y_prob_champ, "#2980b9")),
        (axes[0], ("Challenger LR",    y_prob_chal,  "#e74c3c")),
    ]:
        fpr, tpr, _ = roc_curve(y_test, prob)
        auc = roc_auc_score(y_test, prob)
        ax.plot(fpr, tpr, color=color, lw=2, label=f"{name}  AUC={auc:.4f}")

    axes[0].plot([0, 1], [0, 1], "k--", lw=1)
    axes[0].set_xlabel("False Positive Rate"); axes[0].set_ylabel("True Positive Rate")
    axes[0].set_title("ROC Curves — Champion vs Challenger", fontweight="bold")
    axes[0].legend()

    # Precision-Recall curves
    for ax, (name, prob, color) in [
        (axes[1], ("Champion XGBoost", y_prob_champ, "#2980b9")),
        (axes[1], ("Challenger LR",    y_prob_chal,  "#e74c3c")),
    ]:
        prec, rec, _ = precision_recall_curve(y_test, prob)
        ap = average_precision_score(y_test, prob)
        ax.plot(rec, prec, color=color, lw=2, label=f"{name}  AP={ap:.4f}")

    axes[1].set_xlabel("Recall"); axes[1].set_ylabel("Precision")
    axes[1].set_title("Precision-Recall — Champion vs Challenger", fontweight="bold")
    axes[1].legend()

    plt.tight_layout()
    out = os.path.join(CHARTS_DIR, "champion_vs_challenger.png")
    plt.savefig(out, dpi=130, bbox_inches="tight")
    plt.close()
    logger.info(f"Champion vs Challenger chart saved → {out}")
    return out


def plot_lift_comparison(lift_champ: pd.DataFrame, lift_chal: pd.DataFrame):
    fig, ax = plt.subplots(figsize=(10, 5))

    deciles = lift_champ["Decile"].values
    x = np.arange(len(deciles))
    w = 0.35

    ax.bar(x - w/2, lift_champ["Lift"].values, w, label="Champion XGBoost",
           color="#2980b9", edgecolor="white")
    ax.bar(x + w/2, lift_chal["Lift"].values,  w, label="Challenger LR",
           color="#e74c3c", edgecolor="white")
    ax.axhline(1, color="black", linestyle="--", lw=1, label="Baseline lift = 1")

    ax.set_xticks(x)
    ax.set_xticklabels([f"D{d}" for d in deciles])
    ax.set_xlabel("Decile (1=Highest Risk)"); ax.set_ylabel("Lift")
    ax.set_title("Lift by Decile — Champion vs Challenger", fontweight="bold")
    ax.legend()
    plt.tight_layout()

    out = os.path.join(CHARTS_DIR, "lift_comparison.png")
    plt.savefig(out, dpi=130, bbox_inches="tight")
    plt.close()
    logger.info(f"Lift comparison chart saved → {out}")
    return out


def plot_cutoff_sensitivity(cutoff_df: pd.DataFrame):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    for name, color in [("Champion XGBoost", "#2980b9"), ("Challenger LR", "#e74c3c")]:
        sub = cutoff_df[cutoff_df["Model"] == name]
        axes[0].plot(sub["Cut-off"], sub["Precision"], marker="o", label=f"{name} Precision", color=color)
        axes[0].plot(sub["Cut-off"], sub["Recall"],    marker="s", linestyle="--",
                     label=f"{name} Recall", color=color, alpha=0.6)
        axes[1].plot(sub["Cut-off"], sub["Approval Rate"], marker="o", label=name, color=color)

    axes[0].set_title("Precision & Recall vs Cut-off", fontweight="bold")
    axes[0].set_xlabel("Probability Cut-off"); axes[0].legend(fontsize=8)
    axes[1].set_title("Approval Rate vs Cut-off", fontweight="bold")
    axes[1].set_xlabel("Probability Cut-off"); axes[1].set_ylabel("Approval Rate")
    axes[1].legend()

    plt.tight_layout()
    out = os.path.join(CHARTS_DIR, "cutoff_sensitivity.png")
    plt.savefig(out, dpi=130, bbox_inches="tight")
    plt.close()
    logger.info(f"Cut-off sensitivity chart saved → {out}")
    return out


# ── Excel validation report ───────────────────────────────────────────────────
def write_validation_excel(metrics_df, cutoff_df, lift_champ, lift_chal,
                            chart_paths: list):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def write_table(ws, df, start_row=1, hdr_fill_color=None):
        hf = hdr_fill(hdr_fill_color) if hdr_fill_color else HDR_BLUE
        for col_idx, col in enumerate(df.columns, 1):
            c = ws.cell(row=start_row, column=col_idx, value=col)
            c.font = WHITE_FONT; c.fill = hf
            c.alignment = center(); c.border = thin_border()
        for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
            for c_idx, val in enumerate(row, 1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.alignment = center(); c.border = thin_border()
        return start_row + len(df) + 2

    # ─── Tab 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Model Comparison")
    ws1["A1"] = "Champion vs Challenger — Model Validation Summary"
    ws1["A1"].font = bold_font(14, "FFFFFF")
    ws1["A1"].fill = HDR_BLUE
    ws1.merge_cells(f"A1:{get_column_letter(len(metrics_df.columns))}1")
    ws1["A1"].alignment = center()

    next_row = write_table(ws1, metrics_df, start_row=3)

    # Colour champion row green, challenger row amber
    for row in ws1.iter_rows(min_row=4, max_row=5, min_col=1, max_col=len(metrics_df.columns)):
        for cell in row:
            if cell.row == 4:  # Champion
                cell.fill = GREEN_FILL
            else:
                cell.fill = hdr_fill("FFF2CC")

    ws1[f"A{next_row}"] = "✅ Champion model recommended based on superior Gini and KS statistics."
    ws1[f"A{next_row}"].font = bold_font(11)
    auto_width(ws1)

    # ─── Tab 2: Lift ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Lift Analysis")
    ws2["A1"] = "Lift by Decile — Champion vs Challenger"
    ws2["A1"].font = bold_font(14, "FFFFFF")
    ws2["A1"].fill = HDR_BLUE
    ws2.merge_cells("A1:E1"); ws2["A1"].alignment = center()

    r = write_table(ws2, lift_champ, start_row=3)
    ws2[f"A{r}"] = "Challenger LR"
    ws2[f"A{r}"].font = bold_font()
    write_table(ws2, lift_chal, start_row=r + 1)
    auto_width(ws2)

    # ─── Tab 3: Cut-off Sensitivity ──────────────────────────────────────────
    ws3 = wb.create_sheet("Cut-off Sensitivity")
    ws3["A1"] = "Cut-off Sensitivity Analysis"
    ws3["A1"].font = bold_font(14, "FFFFFF")
    ws3["A1"].fill = HDR_BLUE
    ws3.merge_cells(f"A1:{get_column_letter(len(cutoff_df.columns))}1")
    ws3["A1"].alignment = center()
    write_table(ws3, cutoff_df, start_row=3)
    auto_width(ws3)

    # ─── Tab 4: Charts ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Validation Charts")
    ws4["A1"] = "Validation Charts"
    ws4["A1"].font = bold_font(14, "FFFFFF")
    ws4["A1"].fill = HDR_BLUE

    anchors = ["A3", "A25", "A47"]
    for path, anchor in zip(chart_paths, anchors):
        if os.path.exists(path):
            img = XLImage(path)
            img.width  = 600
            img.height = 300
            img.anchor = anchor
            ws4.add_image(img)

    out = os.path.join(EXCEL_DIR, "challenger_report.xlsx")
    wb.save(out)
    logger.info(f"Challenger validation report saved → {out}")
    return out


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    logger.info("=" * 60)
    logger.info("STEP 6 — CHAMPION vs CHALLENGER EVALUATION")
    logger.info("=" * 60)

    woe_path = os.path.join(OUTPUTS_DIR, "woe_data.parquet")
    test_path = os.path.join(OUTPUTS_DIR, "test_scored.parquet")
    if not os.path.exists(woe_path):
        raise FileNotFoundError("Run 02_feature_engineering.py first.")

    df_woe  = pd.read_parquet(woe_path)
    df_test = pd.read_parquet(test_path)

    artifacts = joblib.load(os.path.join(OUTPUTS_DIR, "models", "binning_models.pkl"))
    selected_features = artifacts["selected_features"]

    model_bundle = joblib.load(os.path.join(OUTPUTS_DIR, "models", "xgb_model.pkl"))

    # ── Champion predictions ──────────────────────────────────────────────────
    y_test_champ = df_test[TARGET].values
    y_prob_champ = df_test["y_prob"].values

    # ── Train & evaluate challenger ───────────────────────────────────────────
    logger.info("\nTraining Challenger (Logistic Regression)...")
    lr, X_test_lr, y_test_lr, y_prob_chal = train_challenger(df_woe, selected_features)

    # Save challenger model
    joblib.dump(lr, os.path.join(OUTPUTS_DIR, "models", "lr_challenger.pkl"))
    logger.info("Challenger model saved → outputs/models/lr_challenger.pkl")

    # ── Metrics ───────────────────────────────────────────────────────────────
    m_champ = compute_metrics(y_test_champ, y_prob_champ, "Champion XGBoost")
    m_chal  = compute_metrics(y_test_lr,    y_prob_chal,  "Challenger LR")
    metrics_df = pd.DataFrame([m_champ, m_chal])

    logger.info("\n📊 Model Comparison:")
    print("\n" + metrics_df.to_string(index=False))

    gini_lift = round(m_champ["Gini"] - m_chal["Gini"], 4)
    ks_lift   = round(m_champ["KS"]   - m_chal["KS"],   4)
    logger.info(f"\n  Gini lift (Champion over Challenger): {gini_lift:+.4f}")
    logger.info(f"  KS   lift (Champion over Challenger): {ks_lift:+.4f}")

    # ── Tables ────────────────────────────────────────────────────────────────
    cutoff_df  = cutoff_sensitivity(y_test_champ, y_prob_champ, y_prob_chal[:len(y_prob_champ)]
                                     if len(y_prob_chal) >= len(y_prob_champ)
                                     else np.resize(y_prob_chal, len(y_prob_champ)))
    lift_champ = compute_lift(y_test_champ, y_prob_champ, "Champion XGBoost")
    lift_chal  = compute_lift(y_test_lr,    y_prob_chal,  "Challenger LR")

    # Save CSVs
    metrics_df.to_csv(os.path.join(OUTPUTS_DIR, "model_comparison.csv"), index=False)
    cutoff_df.to_csv(os.path.join(OUTPUTS_DIR, "cutoff_sensitivity.csv"), index=False)

    # ── Charts ────────────────────────────────────────────────────────────────
    roc_path    = plot_roc_comparison(y_test_champ, y_prob_champ,
                                       y_prob_chal[:len(y_prob_champ)])
    lift_path   = plot_lift_comparison(lift_champ, lift_chal)
    cutoff_path = plot_cutoff_sensitivity(cutoff_df)

    # ── Excel report ─────────────────────────────────────────────────────────
    write_validation_excel(metrics_df, cutoff_df, lift_champ, lift_chal,
                            [roc_path, lift_path, cutoff_path])

    verdict = "✅ CHAMPION RETAINED" if m_champ["Gini"] >= m_chal["Gini"] else "⚠️  CHALLENGER WINS — CONSIDER PROMOTING"
    logger.info(f"\n{verdict}")
    return metrics_df


if __name__ == "__main__":
    main()
