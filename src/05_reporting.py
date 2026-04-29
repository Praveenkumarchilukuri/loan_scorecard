"""
05_reporting.py
Monthly model performance reports in structured Excel format (openpyxl).
Tabs: Score Band Summary | Decile Analysis | Gini Trend | PSI Monitor | Scorecard
Formatted for management consumption and regulatory review.
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import joblib
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                               numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from utils import OUTPUTS_DIR, CHARTS_DIR, EXCEL_DIR, TARGET, gini_coefficient, logger


# ── Style helpers ─────────────────────────────────────────────────────────────
def hdr_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="000000") -> Font:
    return Font(bold=True, size=size, color=color)

def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_BLUE    = hdr_fill("1F3864")
HDR_GREY    = hdr_fill("D9E1F2")
GREEN_FILL  = hdr_fill("C6EFCE")
RED_FILL    = hdr_fill("FFC7CE")
AMBER_FILL  = hdr_fill("FFEB9C")
WHITE_FONT  = bold_font(color="FFFFFF")


def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)


def write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1,
                       header_fill=None, header_font=None):
    """Write a DataFrame to a worksheet with formatting."""
    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font      = header_font or bold_font()
        cell.fill      = header_fill or HDR_GREY
        cell.alignment = center()
        cell.border    = thin_border()

    # Data
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()

    return start_row + len(df) + 1


# ── Score band analysis ───────────────────────────────────────────────────────
def build_score_band_table(df: pd.DataFrame) -> pd.DataFrame:
    bands = [(300, 499, "300–499 (High Risk)"),
             (500, 579, "500–579 (Medium-High)"),
             (580, 619, "580–619 (Medium)"),
             (620, 659, "620–659 (Medium-Low)"),
             (660, 719, "660–719 (Good)"),
             (720, 759, "720–759 (Very Good)"),
             (760, 850, "760–850 (Excellent)")]

    rows = []
    for lo, hi, label in bands:
        mask   = (df["credit_score"] >= lo) & (df["credit_score"] <= hi)
        subset = df[mask]
        n      = len(subset)
        dr     = subset[TARGET].mean() if n > 0 else 0
        rows.append({
            "Score Band":    label,
            "Applicants":    n,
            "% of Book":     round(n / len(df) * 100, 2),
            "Default Rate":  round(dr * 100, 2),
            "Approval Rate": round((1 - dr) * 100, 2),
        })
    return pd.DataFrame(rows)


# ── Decile analysis ───────────────────────────────────────────────────────────
def build_decile_table(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["decile"] = pd.qcut(df["credit_score"], q=10, labels=False, duplicates="drop")
    df["decile"] += 1  # 1=lowest score, 10=highest

    rows = []
    for d in sorted(df["decile"].unique()):
        sub = df[df["decile"] == d]
        rows.append({
            "Decile":           int(d),
            "Min Score":        round(sub["credit_score"].min(), 0),
            "Max Score":        round(sub["credit_score"].max(), 0),
            "Count":            len(sub),
            "Defaults":         int(sub[TARGET].sum()),
            "Default Rate %":   round(sub[TARGET].mean() * 100, 2),
            "Cumulative Lift":  None,  # filled below
        })

    decile_df = pd.DataFrame(rows)
    overall_dr = df[TARGET].mean()
    decile_df["Cumulative Lift"] = (decile_df["Default Rate %"] / (overall_dr * 100)).round(2)
    return decile_df


# ── Gini trend (simulated monthly) ───────────────────────────────────────────
def build_gini_trend(df_test: pd.DataFrame, n_months: int = 6) -> pd.DataFrame:
    rows = []
    rng = np.random.default_rng(0)
    base_gini = gini_coefficient(df_test[TARGET].values, df_test["y_prob"].values)

    for m in range(1, n_months + 1):
        noise = rng.normal(0, 0.008)
        g     = round(base_gini + noise - (m * 0.003), 4)  # slight degradation
        rows.append({"Month": f"Month {m:02d}", "Gini": round(max(g, 0.3), 4)})
    return pd.DataFrame(rows)


def plot_gini_trend(gini_df: pd.DataFrame):
    plt.figure(figsize=(9, 4))
    plt.plot(gini_df["Month"], gini_df["Gini"], marker="o", color="#2980b9",
             linewidth=2, markersize=7)
    plt.axhline(0.4, color="#e74c3c", linestyle="--", lw=1.2, label="Min acceptable Gini = 0.40")
    plt.ylabel("Gini Coefficient"); plt.xlabel("Month")
    plt.title("Monthly Gini Trend — Champion Model", fontweight="bold")
    plt.ylim(0.3, 0.7); plt.legend(); plt.tight_layout()
    out = os.path.join(CHARTS_DIR, "gini_trend.png")
    plt.savefig(out, dpi=130, bbox_inches="tight")
    plt.close()
    logger.info(f"Gini trend chart saved → {out}")


# ── Main report builder ───────────────────────────────────────────────────────
def build_excel_report(df: pd.DataFrame, df_test: pd.DataFrame,
                        scorecard_df: pd.DataFrame, psi_df: pd.DataFrame):

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ─── TAB 1: Cover ────────────────────────────────────────────────────────
    ws_cover = wb.create_sheet("Cover")
    ws_cover.sheet_view.showGridLines = False
    ws_cover.column_dimensions["A"].width = 60

    ws_cover["A1"] = "CREDIT SCORECARD MODEL REPORT"
    ws_cover["A1"].font = Font(bold=True, size=20, color="FFFFFF")
    ws_cover["A1"].fill = hdr_fill("1F3864")
    ws_cover["A1"].alignment = center()
    ws_cover.row_dimensions[1].height = 50

    meta = [
        ("Model Name",       "XGBoost Credit Scorecard v1.0"),
        ("Dataset",          "Give Me Some Credit (Kaggle)"),
        ("Score Range",      "300 – 850"),
        ("Methodology",      "Log-Odds Scaling | PDO=20 | Base=600"),
        ("Report Date",      pd.Timestamp.today().strftime("%d %b %Y")),
        ("Status",           "CHAMPION"),
    ]
    for i, (k, v) in enumerate(meta, 3):
        ws_cover[f"A{i}"] = k
        ws_cover[f"B{i}"] = v
        ws_cover[f"A{i}"].font = bold_font()
        ws_cover[f"A{i}"].fill = HDR_GREY

    ws_cover.column_dimensions["B"].width = 45

    # ─── TAB 2: Score Band Summary ───────────────────────────────────────────
    ws_sb = wb.create_sheet("Score Band Summary")
    ws_sb["A1"] = "Score Band Approval & Default Rate Analysis"
    ws_sb["A1"].font = bold_font(14, "FFFFFF")
    ws_sb["A1"].fill = HDR_BLUE
    ws_sb.merge_cells("A1:E1")
    ws_sb["A1"].alignment = center()

    sb_df = build_score_band_table(df)
    write_df_to_sheet(ws_sb, sb_df, start_row=3, header_fill=HDR_BLUE, header_font=WHITE_FONT)

    # Colour default rate cells
    for row in ws_sb.iter_rows(min_row=4, max_row=4 + len(sb_df) - 1, min_col=4, max_col=4):
        for cell in row:
            val = cell.value
            if isinstance(val, (int, float)):
                cell.fill = GREEN_FILL if val < 10 else (AMBER_FILL if val < 20 else RED_FILL)

    auto_width(ws_sb)

    # ─── TAB 3: Decile Analysis ───────────────────────────────────────────────
    ws_dec = wb.create_sheet("Decile Analysis")
    ws_dec["A1"] = "Default Rate by Score Decile"
    ws_dec["A1"].font = bold_font(14, "FFFFFF")
    ws_dec["A1"].fill = HDR_BLUE
    ws_dec.merge_cells("A1:G1")
    ws_dec["A1"].alignment = center()

    dec_df = build_decile_table(df)
    next_row = write_df_to_sheet(ws_dec, dec_df, start_row=3,
                                  header_fill=HDR_BLUE, header_font=WHITE_FONT)
    auto_width(ws_dec)

    # ─── TAB 4: Gini Trend ───────────────────────────────────────────────────
    ws_gini = wb.create_sheet("Gini Trend")
    ws_gini["A1"] = "Monthly Gini Coefficient Trend"
    ws_gini["A1"].font = bold_font(14, "FFFFFF")
    ws_gini["A1"].fill = HDR_BLUE
    ws_gini.merge_cells("A1:B1")
    ws_gini["A1"].alignment = center()

    gini_df = build_gini_trend(df_test)
    plot_gini_trend(gini_df)
    write_df_to_sheet(ws_gini, gini_df, start_row=3,
                       header_fill=HDR_BLUE, header_font=WHITE_FONT)

    # Embed Gini chart into sheet
    from openpyxl.drawing.image import Image as XLImage
    gini_chart_path = os.path.join(CHARTS_DIR, "gini_trend.png")
    img = XLImage(gini_chart_path)
    img.anchor = "D3"
    ws_gini.add_image(img)
    auto_width(ws_gini)

    # ─── TAB 5: PSI Monitor ──────────────────────────────────────────────────
    ws_psi = wb.create_sheet("PSI Monitor")
    ws_psi["A1"] = "Population Stability Index — Monthly Monitor"
    ws_psi["A1"].font = bold_font(14, "FFFFFF")
    ws_psi["A1"].fill = HDR_BLUE
    ws_psi.merge_cells("A1:F1")
    ws_psi["A1"].alignment = center()

    write_df_to_sheet(ws_psi, psi_df, start_row=3,
                       header_fill=HDR_BLUE, header_font=WHITE_FONT)

    # Colour PSI status
    for row in ws_psi.iter_rows(min_row=4, max_row=3 + len(psi_df), min_col=3, max_col=3):
        for cell in row:
            v = str(cell.value or "")
            if "STABLE" in v:     cell.fill = GREEN_FILL
            elif "MONITOR" in v:  cell.fill = AMBER_FILL
            elif "ACTION" in v:   cell.fill = RED_FILL

    psi_img = XLImage(os.path.join(CHARTS_DIR, "psi_report.png"))
    psi_img.anchor = "H3"
    ws_psi.add_image(psi_img)
    auto_width(ws_psi)

    # ─── TAB 6: Scorecard Points ─────────────────────────────────────────────
    ws_sc = wb.create_sheet("Scorecard Points")
    ws_sc["A1"] = "Scorecard Points Table — Feature Bins & Points Allocation"
    ws_sc["A1"].font = bold_font(14, "FFFFFF")
    ws_sc["A1"].fill = HDR_BLUE
    ws_sc.merge_cells("A1:F1")
    ws_sc["A1"].alignment = center()

    write_df_to_sheet(ws_sc, scorecard_df, start_row=3,
                       header_fill=HDR_BLUE, header_font=WHITE_FONT)
    auto_width(ws_sc)

    # ─── Save ────────────────────────────────────────────────────────────────
    out_path = os.path.join(EXCEL_DIR, "model_report.xlsx")
    wb.save(out_path)
    logger.info(f"Excel model report saved → {out_path}")
    return out_path


def main():
    logger.info("=" * 60)
    logger.info("STEP 5 — EXCEL REPORTING")
    logger.info("=" * 60)

    for f, name in [
        (os.path.join(OUTPUTS_DIR, "scored_data.parquet"),    "scored_data"),
        (os.path.join(OUTPUTS_DIR, "test_scored.parquet"),    "test_scored"),
        (os.path.join(OUTPUTS_DIR, "scorecard_points.csv"),   "scorecard"),
        (os.path.join(OUTPUTS_DIR, "psi_results.csv"),        "psi_results"),
    ]:
        if not os.path.exists(f):
            raise FileNotFoundError(f"Missing {name}. Run previous steps first.")

    df          = pd.read_parquet(os.path.join(OUTPUTS_DIR, "scored_data.parquet"))
    df_test     = pd.read_parquet(os.path.join(OUTPUTS_DIR, "test_scored.parquet"))
    scorecard_df = pd.read_csv(os.path.join(OUTPUTS_DIR, "scorecard_points.csv"))
    psi_df      = pd.read_csv(os.path.join(OUTPUTS_DIR, "psi_results.csv"))

    out = build_excel_report(df, df_test, scorecard_df, psi_df)
    logger.info(f"\n✅  Report ready: {out}")


if __name__ == "__main__":
    main()
